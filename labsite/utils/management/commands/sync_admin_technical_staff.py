import html
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from labsite.standardpages.models import FacultyPage, StandardPage


class Command(BaseCommand):
    help = "Sync pharmacy admin and center staff from Excel to the admin/technical staff page."

    def add_arguments(self, parser):
        parser.add_argument(
            "xlsx_path",
            nargs="?",
            default="../0_important_data/国重室成员名单-孟山水revised.xlsx",
            help="Excel file containing the '药学院管理及中心人员' sheet.",
        )
        parser.add_argument(
            "--keep-faculty-pages",
            action="store_true",
            help="Do not delete matching FacultyPage records.",
        )

    def handle(self, *args, **options):
        xlsx_path = Path(options["xlsx_path"])
        if not xlsx_path.exists():
            raise CommandError(f"Excel file does not exist: {xlsx_path}")

        rows = self.read_staff_rows(xlsx_path)
        if not rows:
            raise CommandError("No staff rows found in sheet: 药学院管理及中心人员")

        try:
            page = StandardPage.objects.get(slug="admin-technical-staff")
        except StandardPage.DoesNotExist as exc:
            raise CommandError("Page not found: admin-technical-staff") from exc

        names = [row["name"] for row in rows]
        with transaction.atomic():
            deleted_count = 0
            if not options["keep_faculty_pages"]:
                for faculty_page in FacultyPage.objects.filter(title__in=names).specific():
                    faculty_page.delete()
                    deleted_count += 1

            page.body = [("html", self.staff_table_html(rows))]
            page.save_revision().publish()

        self.stdout.write(
            self.style.SUCCESS(
                f"Synced {len(rows)} staff records to {page.url}; deleted {deleted_count} faculty pages."
            )
        )

    def read_staff_rows(self, xlsx_path):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl is required to read Excel files.") from exc

        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
        if "药学院管理及中心人员" not in workbook.sheetnames:
            raise CommandError("Sheet not found: 药学院管理及中心人员")

        sheet = workbook["药学院管理及中心人员"]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = self.clean_cell(row[0] if len(row) > 0 else "")
            email = self.clean_cell(row[1] if len(row) > 1 else "")
            if not name:
                continue
            rows.append(
                {
                    "index": len(rows) + 1,
                    "name": name,
                    "email": email,
                    "unit": "中山大学药学院",
                }
            )
        return rows

    def clean_cell(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def staff_table_html(self, rows):
        table_rows = []
        for row in rows:
            email_html = html.escape(row["email"])
            if row["email"]:
                email_html = f'<a href="mailto:{email_html}">{email_html}</a>'
            table_rows.append(
                "<tr>"
                f"<td>{row['index']}</td>"
                f"<td>{html.escape(row['name'])}</td>"
                f"<td>{email_html}</td>"
                f"<td>{html.escape(row['unit'])}</td>"
                "</tr>"
            )

        return (
            '<div class="staff-table-block">'
            "<p>药学院管理及中心人员为实验室运行提供行政管理、公共平台与技术支撑服务。</p>"
            '<div class="staff-table-scroll">'
            '<table class="staff-table">'
            "<thead><tr><th>序号</th><th>姓名</th><th>邮箱</th><th>单位</th></tr></thead>"
            f"<tbody>{''.join(table_rows)}</tbody>"
            "</table>"
            "</div>"
            "</div>"
        )
